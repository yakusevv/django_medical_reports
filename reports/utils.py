from docxtpl import DocxTemplate, InlineImage
from docx.shared import Mm
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from django.utils.translation import ugettext as _
from django.db.models import Min, Max

from .models import ReportTemplate


def docx_report_generator(report, type_of_report):
    try:
        doc_path = ReportTemplate.objects.get(
            country=report.city.district.region.country
            ).template
        doc = DocxTemplate(doc_path)
        images = []
        for image in report.additional_images.get_queryset():
            if image.expand:
                images.append(InlineImage(doc, image.image, width=Mm(170)))
            else:
                images.append(InlineImage(doc, image.image, width=Mm(110)))

        context = {
                'R': report,
                'i': images,
                'd': report.diagnosis.get_queryset(),
                's': report.service_items.get_queryset(),
                't': type_of_report
                }
        doc.render(context)

        return doc
    except ReportTemplate.DoesNotExist:
        return None


def reports_xlsx_generator(reports):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Reports'

    columns = [
        _('ref. number'),
        _('Company ref. number'),
        _('Full name, date of birth'),
        _('City'),
        _('Total price\nfor the doctor'),
        _('Total\nprice')
    ]

    row_num = 2
    col_first = 2

    for col_num, column_title in enumerate(columns, col_first):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.border = Border(
                        bottom=Side(border_style='medium', color='FF000000'),
                        top=Side(border_style='medium', color='FF000000'),
                        left=Side(border_style='medium', color='FF000000'),
                        right=Side(border_style='medium', color='FF000000'),
                    )
        cell.fill = PatternFill(
                        start_color='f7f7f9',
                        end_color='f7f7f9',
                        fill_type='solid',
                    )
        cell.alignment = Alignment(wrapText=True)
        col_letter = get_column_letter(col_num)
        if col_letter == 'B':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 20
        elif col_letter == 'C':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 25
        elif col_letter == 'D':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 35
        elif col_letter == 'E':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 20
        elif col_letter == 'F':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 15
        elif col_letter == 'G':
            column_dimensions = worksheet.column_dimensions[col_letter]
            column_dimensions.width = 15

    for index, report in enumerate(reports):
        row_num += 1

        row = [
            report.get_full_ref_number,
            report.company_ref_number,
            ' '.join((
                    str(report.patients_last_name),
                    str(report.patients_first_name)
                    )) + ', ' + str(report.patients_date_of_birth.strftime("%d.%m.%Y")),
            report.city.name,
            report.get_total_price_doctor,
            report.get_total_price
        ]

        for col_num, cell_value in enumerate(row, col_first):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.number_format = '0.00'
            cell.value = cell_value
            if index == len(reports) - 1:
                if col_num == col_first:
                    cell.border = Border(
                        left=Side(border_style='medium', color='FF000000'),
                        bottom=Side(border_style='medium', color='FF000000')
                    )
                elif col_num == (len(row) - 1) + col_first:
                    cell.border = Border(
                        right=Side(border_style='medium', color='FF000000'),
                        bottom=Side(border_style='medium', color='FF000000')
                    )
                else:
                    cell.border = Border(
                        left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='medium', color='FF000000')
                    )
            else:
                if col_num == col_first:
                    cell.border = Border(
                        left=Side(border_style='medium', color='FF000000'),
                    )
                elif col_num == (len(row) - 1) + col_first:
                    cell.border = Border(
                        right=Side(border_style='medium', color='FF000000'),
                    )
                else:
                    cell.border = Border(
                        left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000')
                    )

    return workbook


def reports_xlsx_generator_new_format(reports, type_of_table):

    def table(queryset, year):

        columns = [
            _('ref. number'),
            _('Company ref. number'),
            _('Full name, date of birth'),
            _('City'),
            _('Total price for the doctor'),
        ]
        if type_of_table == "staff":
            columns.append(_('Total price'))

        data = []

        for report in queryset:

            row = [
                report.get_full_ref_number,
                report.get_full_company_ref_number,
                ' '.join((
                    str(report.patients_last_name),
                    str(report.patients_first_name)
                )) + ', ' + str(report.patients_date_of_birth.strftime("%d.%m.%Y")),
                report.city.name,
                report.get_total_price_doctor,
            ]
            if type_of_table == "staff":
                row.append(report.get_total_price)

            data.append(row)

        worksheet.append(columns)

        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 25
        worksheet.column_dimensions["C"].width = 45
        worksheet.column_dimensions["D"].width = 25
        worksheet.column_dimensions["E"].width = 25

        if type_of_table == "staff":
            worksheet.column_dimensions["F"].width = 25
            table_ref = "A1:F"
        else:
            table_ref = "A1:E"
        for row in data:
            worksheet.append(row)

        tab = Table(displayName='year_' + str(year), ref=table_ref + str(len(data) + 1), totalsRowShown=True)

        style = TableStyleInfo(
                               name="TableStyleMedium9",
                               showFirstColumn=False,
                               showLastColumn=False,
                               showRowStripes=True,
                               showColumnStripes=False
                               )
        tab.tableStyleInfo = style

        worksheet.add_table(tab)

    workbook = Workbook()
    worksheet = workbook.active
    if reports:
        workbook.remove_sheet(worksheet)
        years_range = range(
                        reports.aggregate(Min('report_request__date_time'))['report_request__date_time__min'].year,
                        reports.aggregate(Max('report_request__date_time'))['report_request__date_time__max'].year + 1
                        )
        years = [year for year in years_range]
        for year in years:
            worksheet = workbook.create_sheet(str(year))
            queryset_year = reports.filter(report_request__date_time__year=year)
            table(queryset_year, year)

    return workbook

